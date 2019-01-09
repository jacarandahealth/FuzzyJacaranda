import csv
import os
import sys
from string import ascii_lowercase
from time import sleep

import PySimpleGUI as sg
from fuzzywuzzy import fuzz
from openpyxl import load_workbook


class Fuzzy:

    def __init__(self):
        pass

    def gui(self):
        desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        layout = [
            [sg.Text('JHSL Facility Fuzzy Match', size=(45, 1), font=("Helvetica", 15), text_color='green',
                     justification='Center')],
            [sg.Text('==============================================================================================')],
            [sg.Text('Select the Excel file you want to process (Input)', size=(55, 1), font=("Helvetica", 15),
                     text_color='blue4')],
            [sg.InputText(), sg.FileBrowse()],
            [sg.Text('')],
            [sg.Text('Enter the column letter containing the raw location (facility name)',
                     size=(70, 1), font=("Helvetica", 15), text_color='blue4')],
            [sg.Text('          ...and the column letter where the output will be inserted',
                     size=(70, 1), font=("Helvetica", 15), text_color='blue4')],
            [sg.Text('Raw Location Column', size=(25, 1)), sg.InputText()],
            [sg.Text('Fuzzy Match Output Column', size=(25, 1)), sg.InputText()],
            [sg.Text('')],
            [sg.Text('Select the Output Folder', size=(35, 1), font=("Helvetica", 15), text_color='blue4')],
            [sg.InputText(desktop), sg.FolderBrowse()],
            [sg.Text('==============================================================================================')],
            [sg.Submit(), sg.Cancel()]]

        window = sg.Window('FUZZY MATCHING ASSISTANT').Layout(layout)
        event, values = window.Read()

        if event is 'Cancel':
            sys.exit(0)
        else:
            window.Close()
            return values

    # fuzzy match raw location against (modified) DHIS2 facility list
    def fuzzy_match(self, input_values):
        input_filepath = input_values[0]
        raw_location_column_char = input_values[1].lower()
        fuzzy_match_values_col = input_values[2].lower()
        fuzzy_match_values_col_num = ascii_lowercase.index(fuzzy_match_values_col) + 1
        output_directory = input_values[3]
        raw_location_column_num = ascii_lowercase.index(raw_location_column_char) + 1
        input_filename = input_filepath.split('/')[-1].replace('.xlsx', '')
        output_filename = '{}_{}.xlsx'.format(input_filename, 'fuzzy_matched')
        output_filepath = os.path.join(output_directory, output_filename)
        exclude_list = ['none', 'home', 'house', 'nyumbani', 'bmama', 'bmimba', 'm2mama', 'm2mimba', 'mama', 'mimba']

        wb = load_workbook(input_filepath)
        ws = wb.active

        facility_file = 'C:\\JHSL\\FuzzyMatch\\parameters\\main_facility_list.csv'
        with open(facility_file, mode='r') as infile:
            reader = csv.reader(infile)
            facility_dict = dict(reader)

        best_match_dict = {}
        temp_match_dict = {}

        max_row = len(ws['A'])
        counter = 1

        raw_location_exclude_list = ['yes', 'ndio', 'hospital', 'hospitali', 'centre', 'center', 'services', 'clinic']
        # 'sub', 'district']

        ws.insert_cols(fuzzy_match_values_col_num)  # insert column and write results to this column
        # iterate through worksheets, match each raw_location against entire facility list, write the best match if match ratio >= 10
        for each_row in ws.iter_rows(min_row=2, min_col=raw_location_column_num,
                                     max_col=raw_location_column_num):  # iterate through col C

            sg.OneLineProgressMeter('Fuzzy Match Progress Meter', counter, max_row - 1, 'key')

            raw_loc = each_row[0].value

            try:
                raw_location_lower = raw_loc.lower()
            except TypeError:
                raw_location_lower = raw_loc
            except AttributeError:
                raw_location_lower = ''

            for word in raw_location_exclude_list:
                if word in raw_location_lower:
                    raw_location_lower = raw_location_lower.replace(word, '')

            # eliminate strings which are too long and too short
            if len(raw_location_lower) > 3 and len(raw_location_lower) <= 45:
                if raw_location_lower is not None and raw_location_lower is not '--':

                    # the statement below is to counteract weird matching results, will have a look at that functionality later
                    if raw_location_lower == 'chwele':
                        raw_location_lower = 'chwele sub district hospital'
                    elif raw_location_lower == 'chwele hospital':
                        raw_location_lower = 'chwele sub district hospital'
                    elif raw_location_lower == 'lugulu' or raw_location_lower == 'lukulu':
                        raw_location_lower = 'lugulu friends mission hospital'
                    elif raw_location_lower == 'thika' or raw_location_lower == 'thika level 5 hospital' or raw_location_lower == 'level 5 thika':
                        raw_location_lower = 'thika level 5'
                    elif raw_location_lower == 'bungoma hospital':
                        raw_location_lower = 'bungoma referral'
                    elif raw_location_lower == 'kiandutu':
                        raw_location_lower = 'kiandutu health centre'
                    elif raw_location_lower == 'kiambu' or raw_location_lower == 'kiambu hospital':
                        raw_location_lower = 'kiandutu health centre'
                    elif raw_location_lower == 'tigoni':
                        raw_location_lower = 'tigoni district hospital'
                    elif raw_location_lower == 'bungoma hospital':
                        raw_location_lower = 'bungoma county referral hospital'
                    elif raw_location_lower == 'kihara':
                        raw_location_lower = 'kihara sub-county hospital'
                    elif raw_location_lower == 'kihara':
                        raw_location_lower = 'kihara sub-county  hospital'
                    elif 'mama' in raw_location_lower and 'lucy' not in raw_location_lower and 'margaret' not in raw_location_lower:
                        raw_location_lower = ''
                    elif raw_location_lower == 'mother':
                        raw_location_lower = ''
                    elif raw_location_lower in exclude_list:
                        raw_location_lower = ''

                    for each_facility in facility_dict.keys():  # iterate through entire facility file
                        fuzz_ratio = fuzz.partial_ratio(raw_location_lower,
                                                        each_facility)  # get partial match ratio
                        temp_match_dict[each_facility] = fuzz_ratio  # facility: ratio

                    # get highest value in temp_match_dict and return key
                    # https://stackoverflow.com/questions/268272/getting-key-with-maximum-value-in-dictionary
                    best_match = max(temp_match_dict, key=temp_match_dict.get)

                    best_match_dict[best_match] = fuzz_ratio  # add best match to dict

                if best_match_dict[best_match] >= 10:  # only add matches with a ratio greater than 10
                    cell_index = fuzzy_match_values_col + str(each_row[0].row)
                    ws[cell_index] = facility_dict[best_match]  # + " " + str(fuzz_ratio)  # write to cell

                # print('{}: {}: {}'.format(raw_location_lower, best_match, fuzz_ratio))  # leave for debugging
            counter += 1
        ws[fuzzy_match_values_col + '1'] = 'Fuzzy Match Results'  # Results column header
        wb.save(output_filepath)


class Main():

    def __init__(self):
        pass

    def run(self):
        fuzz = Fuzzy()
        input_values = fuzz.gui()
        fuzz.fuzzy_match(input_values)
        print('Complete!')
        print('This window will self-destruct in 10 seconds!')
        sleep(10)
        sys.exit(0)


if __name__ == '__main__':
    main = Main()
    main.run()
